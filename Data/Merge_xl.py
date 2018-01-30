from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
import geocoder

wb_new = xlsxwriter.Workbook("UberFinal2.xlsx")
#ws = wb_new.create_sheet()
trips_ws = wb_new.add_worksheet("UberData") #Open Existing workbook

    #Dformat = log_book.add_format()
bold = wb_new.add_format({'bold': True})

trips_ws.set_column('A:A',20)
trips_ws.set_column('F:G',60)
trips_ws.set_column('H:K',25)
trips_ws.set_row(0,15,bold)
trips_ws.write('A1', 'Date', bold)
trips_ws.write('B1', "Time",bold)
trips_ws.write('C1', "Fare")
trips_ws.write('D1', "Duration")
trips_ws.write('E1', "Distance")
trips_ws.write('F1', "PICKUP")
trips_ws.write('G1', "DROP")
trips_ws.write('H1', "PICKUP LATITUDE")
trips_ws.write('I1', "PICKUP LONGITUDE")
trips_ws.write('J1', "DROP LONGITUDE")
trips_ws.write('K1', "DROP LONGITUDE")


r =0
for m in range(1,13):
    print "\n"
    if m<10:
        mon = "0"+str(m)
    else:
        mon = str(m)
        status = 0
    for d in range(1,32):
        if d<10:
            day = "0"+str(d)
        else:
            day = str(d)
        date = mon+"_"+day+"_"+"2017"
        try:
            filename = "UberData_"+ date + ".xlsx"
            wb = load_workbook(filename, read_only=True)    #OPENING WORKBOOK FOR READING
            status = 1
        except:
            status = 0

        if status == 1:    #FILE FOUND
            ws = wb['UberData']
            strt =1
            for row in ws.rows:              #READING WORKBOOK
                if strt ==1:
                    strt+=1
                else:
                    try:
                        if row[4].value==0 :
                            pass
                        else:               #WRITING DATA TO ANOTHER WORKBOOK
                            trips_ws.write('A' + str(r + 2), row[0].value)  # ADD DATE
                            trips_ws.write('B' + str(r + 2), row[1].value)  # ADD TIME
                            trips_ws.write_number('C' + str(r + 2), row[2].value)  # ADD FARE
                            trips_ws.write_number('D' + str(r + 2), row[3].value)  # ADD TRIP DURATION
                            trips_ws.write_number('E' + str(r + 2), row[4].value)  # ADD DISTANCE
                            trips_ws.write_string('F' + str(r + 2), row[5].value)  # ADD PICKUP
                            trips_ws.write_string('G' + str(r + 2), row[6].value)  # ADD DROP
                            gp = geocoder.google(row[5].value)
                            trips_ws.write_number('H' + str(r + 2), gp.lat)  # ADD DROP
                            trips_ws.write_number('I' + str(r + 2), gp.lng)  # ADD DROP
                            gd = geocoder.google(row[6].value)
                            trips_ws.write_number('J' + str(r + 2), gd.lat)  # ADD DROP
                            trips_ws.write_number('K' + str(r + 2), gd.lng)  # ADD DROP

                            r = r+1
                        strt+=1
                    except Exception as e:
                        print "Error : ",e
wb_new.close()









